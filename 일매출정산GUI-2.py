import os
import sqlite3
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
import re
import sys
import subprocess

# 데이터베이스 파일명 및 엑셀 파일명
DB_FILENAME = '카페5월5일매출.db'
EXCEL_FILENAME = '보고서출력.xlsx'

# 기본 설정
root = tk.Tk()
root.title('카페 매출 입력')
font = ('맑은고딕', 11)

# 수량 입력란을 비활성화할 항목 리스트
disable_quantity_items = ["현금", "카드", "인터넷"]

# 금액 입력란을 비활성화할 항목
disable_amount_item = "직원제공"

# 날짜 관련 변수 및 함수
current_date = datetime.now() - timedelta(days=1)
date_var = tk.StringVar(value=current_date.strftime('%Y-%m-%d'))

def update_date(days):
    global current_date
    current_date += timedelta(days=days)
    date_var.set(current_date.strftime('%Y-%m-%d'))
    load_existing_data()

def format_date(event=None):
    raw_date = date_var.get()
    if len(raw_date) == 8 and raw_date.isdigit():
        formatted = f"{raw_date[:4]}-{raw_date[4:6]}-{raw_date[6:]}"
        date_var.set(formatted)
    load_existing_data()

# 상품 리스트 가져오기
def fetch_product_list():
    conn = sqlite3.connect(DB_FILENAME)
    cursor = conn.cursor()
    cursor.execute("SELECT 상품명 FROM 상품리스트 ORDER BY No ASC")
    products = [row[0] for row in cursor.fetchall()]
    conn.close()
    return products

product_names = fetch_product_list()

# 수식 평가 함수
def evaluate_expression(expression):
    try:
        sanitized_expression = re.sub(r'[^0-9+\-*/().]', '', expression)
        return eval(sanitized_expression)
    except Exception:
        return expression

# 기존 데이터 로드 함수
def load_existing_data():
    date = date_var.get()
    conn = sqlite3.connect(DB_FILENAME)
    cursor = conn.cursor()
    
    for product_name, quantity_entry, amount_entry in entries:
        cursor.execute('''
            SELECT 수량, 금액 FROM 매출
            WHERE 날짜 = ? AND 상품ID = (SELECT id FROM 상품리스트 WHERE 상품명 = ?)
        ''', (date, product_name))
        result = cursor.fetchone()
        
        if product_name not in disable_quantity_items:
            quantity_entry.delete(0, tk.END)
        if product_name != disable_amount_item:
            amount_entry.delete(0, tk.END)
        
        if result:
            if product_name not in disable_quantity_items:
                quantity_entry.insert(0, str(result[0]))
            if product_name != disable_amount_item:
                amount_entry.insert(0, str(result[1]))
    
    conn.close()

# DB에 데이터 삽입 함수
def save_to_db():
    conn = sqlite3.connect(DB_FILENAME)
    cursor = conn.cursor()
    
    date = date_var.get()
    
    cursor.execute("DELETE FROM 매출 WHERE 날짜 = ?", (date,))
    
    for product_name, quantity_entry, amount_entry in entries:
        quantity = quantity_entry.get().strip()
        amount = amount_entry.get().strip()
        
        quantity = int(quantity) if quantity.isdigit() else 0
        amount = int(amount) if amount.isdigit() else 0
        
        if quantity != 0 or amount != 0:
            cursor.execute('''
                INSERT INTO 매출 (날짜, 상품ID, 수량, 금액) 
                VALUES (?, (SELECT id FROM 상품리스트 WHERE 상품명=?), ?, ?)
            ''', (date, product_name, quantity, amount))
    
    conn.commit()
    conn.close()
    messagebox.showinfo("성공", "데이터가 성공적으로 저장되었습니다.")
    
    update_excel_cells()
    write_sales_to_excel(date)
    open_excel_file()
    
    root.quit()
    sys.exit()

# 엑셀 파일 업데이트 함수 (초기화용)
def fetch_cell_positions():
    conn = sqlite3.connect(DB_FILENAME)
    cursor = conn.cursor()
    cursor.execute("SELECT 수량_셀, 금액_셀 FROM 상품리스트")
    cell_positions = cursor.fetchall()
    conn.close()
    return cell_positions

def update_excel_cells():
    cell_positions = fetch_cell_positions()
    wb = load_workbook(EXCEL_FILENAME)
    sheet = wb.active

    for quantity_cell, amount_cell in cell_positions:
        if quantity_cell:
            sheet[quantity_cell].value = 0
        if amount_cell:
            sheet[amount_cell].value = 0

    wb.save(EXCEL_FILENAME)

# 매출 데이터를 엑셀에 작성하는 함수
def write_sales_to_excel(date):
    conn = sqlite3.connect(DB_FILENAME)
    cursor = conn.cursor()

    query = '''
        SELECT m.수량, m.금액, p.수량_셀, p.금액_셀 
        FROM 매출 m 
        JOIN 상품리스트 p ON m.상품ID = p.id 
        WHERE m.날짜 = ?
    '''
    cursor.execute(query, (date,))
    results = cursor.fetchall()

    # 월별 합계 쿼리
    month_start = f"{date[:7]}-01"
    summary_query = '''
        SELECT 
            SUM(CASE WHEN p.id LIKE 'ame' THEN m.수량 ELSE 0 END) as ame_quantity,
            SUM(CASE WHEN p.id LIKE 'ame' THEN m.금액 ELSE 0 END) as ame_amount,
            SUM(CASE WHEN p.id LIKE 'etc_coffee' THEN m.수량 ELSE 0 END) as etc_coffee_quantity,
            SUM(CASE WHEN p.id LIKE 'etc_coffee' THEN m.금액 ELSE 0 END) as etc_coffee_amount,
            SUM(CASE WHEN p.id LIKE 'etc_beverage' THEN m.수량 ELSE 0 END) as etc_beverage_quantity,
            SUM(CASE WHEN p.id LIKE 'etc_beverage' THEN m.금액 ELSE 0 END) as etc_beverage_amount,
            SUM(CASE WHEN p.id LIKE 'desert' THEN m.수량 ELSE 0 END) as desert_quantity,
            SUM(CASE WHEN p.id LIKE 'desert' THEN m.금액 ELSE 0 END) as desert_amount,
            SUM(CASE WHEN p.id LIKE 'place' THEN m.수량 ELSE 0 END) as place_quantity,
            SUM(CASE WHEN p.id LIKE 'place' THEN m.금액 ELSE 0 END) as place_amount,
            SUM(CASE WHEN p.id LIKE 'employees_supply' THEN m.수량 ELSE 0 END) as employees_supply_quantity,
            SUM(CASE WHEN p.id LIKE 'etc' THEN m.수량 ELSE 0 END) as etc_quantity,
            SUM(CASE WHEN p.id LIKE 'etc' THEN m.금액 ELSE 0 END) as etc_amount,
            SUM(CASE WHEN p.id IN ('event_in', 'etc_in') THEN m.금액 ELSE 0 END) as event_etc_in_amount
        FROM 매출 m 
        JOIN 상품리스트 p ON m.상품ID = p.id 
        WHERE m.날짜 BETWEEN ? AND ?
    '''
    cursor.execute(summary_query, (month_start, date))
    summary_results = cursor.fetchone()

    wb = load_workbook(EXCEL_FILENAME)
    sheet = wb.active

    # D6 셀에 date_var 값 입력
    sheet['D6'] = date_var.get()

    # 기존 데이터 입력
    for quantity, amount, quantity_cell, amount_cell in results:
        if quantity_cell:
            sheet[quantity_cell].value = quantity
        if amount_cell:
            sheet[amount_cell].value = amount

    # 새로운 요약 데이터 입력
    sheet['D31'] = summary_results[0]  # ame 수량
    sheet['E31'] = summary_results[1]  # ame 금액
    sheet['D32'] = summary_results[2]  # etc_coffee 수량
    sheet['E32'] = summary_results[3]  # etc_coffee 금액
    sheet['D33'] = summary_results[4]  # etc_beverage 수량
    sheet['E33'] = summary_results[5]  # etc_beverage 금액
    sheet['D34'] = summary_results[6]  # desert 수량
    sheet['E34'] = summary_results[7]  # desert 금액
    sheet['D35'] = summary_results[8]  # place 수량
    sheet['E35'] = summary_results[9]  # place 금액
    sheet['D36'] = summary_results[10]  # employees_supply 수량
    sheet['D37'] = summary_results[11]  # etc 수량
    sheet['E37'] = summary_results[12]  # etc 금액
    sheet['I37'] = summary_results[13]  # event_in + etc_in 금액

    wb.save(EXCEL_FILENAME)
    conn.close()

# 엑셀 파일 열기 함수
def open_excel_file():
    try:
        subprocess.Popen(['start', EXCEL_FILENAME], shell=True)
        print(f"엑셀 파일 '{EXCEL_FILENAME}'이 열렸습니다.")
    except Exception as e:
        print(f"엑셀 파일을 여는 중 오류가 발생했습니다: {e}")

# GUI 구성
frame_date = ttk.Frame(root)
frame_date.pack(pady=10)

btn_prev = ttk.Button(frame_date, text='<<', command=lambda: update_date(-1))
btn_prev.pack(side='left')

entry_date = ttk.Entry(frame_date, textvariable=date_var, font=font, width=12)
entry_date.pack(side='left', padx=5)
entry_date.bind('<FocusOut>', format_date)

btn_next = ttk.Button(frame_date, text='>>', command=lambda: update_date(1))
btn_next.pack(side='left')

frame_products = ttk.Frame(root)
frame_products.pack(pady=10)

ttk.Label(frame_products, text='품명', font=font).grid(row=0, column=0, padx=5)
ttk.Label(frame_products, text='수량', font=font).grid(row=0, column=1, padx=5)
ttk.Label(frame_products, text='금액', font=font).grid(row=0, column=2, padx=5)

entries = []
for i, product_name in enumerate(product_names, start=1):
    ttk.Label(frame_products, text=product_name, font=font).grid(row=i, column=0, padx=5)
    
    entry_quantity = ttk.Entry(frame_products, font=font, width=10, justify='right')
    entry_quantity.grid(row=i, column=1, padx=5)
    
    entry_amount = ttk.Entry(frame_products, font=font, width=10, justify='right')
    entry_amount.grid(row=i, column=2, padx=5)

    def on_focus_out(event=None, entry_field=None):
        value = entry_field.get()
        entry_field.delete(0, tk.END)
        entry_field.insert(0, str(evaluate_expression(value)))
    
    entry_quantity.bind('<FocusOut>', lambda e, field=entry_quantity: on_focus_out(e, field))
    entry_amount.bind('<FocusOut>', lambda e, field=entry_amount: on_focus_out(e, field))
    
    if product_name in disable_quantity_items:
        entry_quantity.config(state='disabled')
        entry_quantity.insert(0, '0')
    
    if product_name == disable_amount_item:
        entry_amount.config(state='disabled')
        entry_amount.insert(0, '0')

    entries.append((product_name, entry_quantity, entry_amount))

# 입력 버튼
btn_save = ttk.Button(root, text='입력', command=save_to_db)
btn_save.pack(pady=10)

root.mainloop()
